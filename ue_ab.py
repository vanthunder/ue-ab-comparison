"""
Unreal Engine A/B Comparison Excel Generator (Post‑Processing Tool)
==================================================================

This script takes aggregated metric workbooks (produced by a preceding
"Performance Analysis Tool for Unreal Engine CSV Exports") and builds a
research‑grade A/B comparison Excel report with per‑scene sheets and a summary.

UPSTREAM (PRECEDING) TOOL (Context)
----------------------------------
The earlier pipeline step (referenced here for clarity) processes raw profiler
CSV exports and produces aggregated per‑scene/variant workbooks.
Summary of that tool:

    Performance Analysis Tool for Unreal Engine CSV Exports
    Robust CSV → Excel Report Generator

    Author: Marvin Schubert
    Version: 1.0.0
    Date: September 2025

    INSTALLATION REQUIREMENTS:
        pip install pandas numpy openpyxl

    USAGE (upstream):
        1. Create folder 'messungen' beside the script
        2. Put EXP_*.csv files inside (pattern: EXP_[Scene]_[A|B]_Messung_[Run].csv)
        3. Run: python messung_auswertung.py
        4. Get an Excel aggregation (p95 / mean metrics per scene & variant)

You then feed the produced workbook(s) into THIS A/B comparison tool.

WHAT THIS TOOL DOES
-------------------
Creates per‑scene aggregated tables for variants A and B and a comparison table
with relative percent deltas (Δ = (B − A) / A · 100). Handles either a single
combined workbook or two variant‑specific workbooks. Adds a global summary.

INSTALLATION (this tool)
------------------------
    pip install -r requirements.txt
  or explicitly:
    pip install pandas numpy openpyxl

WORKFLOW PIPELINE
-----------------
1. Run upstream CSV → aggregation tool to obtain:
       - Single combined:  messungen_auswertung.xlsx
         (contains sheets "Scene{n}_A" & "Scene{n}_B")
       OR
       - Two files:        messungen_auswertung_a.xlsx / messungen_auswertung_b.xlsx
         (sheets may be named "Scene{n}" or already suffixed with _A / _B)
2. Run this script (examples below) to produce the A/B comparison workbook:
       messungen_ab_vergleich.xlsx (default name)

INPUT MODES
-----------
1) Single workbook: 'messungen_auswertung.xlsx' with sheets "Scene{n}_A" / "Scene{n}_B".
2) Two workbooks: 'messungen_auswertung_a.xlsx' & 'messungen_auswertung_b.xlsx' (or via --a/--b) with sheets "Scene{n}" or suffixed.

PER SHEET (INPUT) FORMAT
------------------------
Column 1: metric labels (German labels intentionally retained to match source export)
Columns 2..k: numeric run values (German or English numeric formatting tolerated: '.' thousands + ',' decimal; multi‑dot thousand grouping supported).

OUTPUT
------
Workbook 'messungen_ab_vergleich.xlsx' containing, per scene:
    Scene{n}_Agg_A         (Aggregated metrics variant A)
    Scene{n}_Agg_B         (Aggregated metrics variant B)
    Scene{n}_Vergleich     (Tabular comparison with Δ B vs A [%])
Plus sheet: 'Gesamtübersicht' (summary across all scenes).

METHODOLOGICAL NOTES
--------------------
* Means: Arithmetic mean across run‑level aggregated values (p95 values are already per‑run aggregates upstream).
* Optional FPS recomputation (default ON): Derives FPS per run from mean frametime (1000 / ms) and averages those values (avoids bias of direct FPS arithmetic mean).
* Delta formula: Δ = (B − A) / A · 100, only when A is finite and non‑zero.
* Formatting: German style decimal comma, NBSP thousands; integer counters (N, Draw Calls, Primitives) have 0 decimals; times, memory, FPS, Δ use 3 decimals.

CLI QUICK START
---------------
    # Single combined workbook mode
    python ue_ab.py

    # Two explicit workbooks
    python ue_ab.py --a messungen_auswertung_a.xlsx --b messungen_auswertung_b.xlsx

    # Auto-detect *_a.xlsx / *_b.xlsx in current directory
    python ue_ab.py --auto

    # Custom output filename
    python ue_ab.py --auto --out vergleich_report.xlsx

    # Keep original FPS (skip recomputation)
    python ue_ab.py --no-recompute-fps --auto

ARGUMENTS
---------
--a / --b            Paths to variant A/B workbooks
--auto               Auto-detect default *_a.xlsx / *_b.xlsx if present
--out PATH           Output Excel file name/path (default: messungen_ab_vergleich.xlsx)
--no-recompute-fps   Disable FPS recomputation from frametime means
--debug              Print per-scene parsing and missing-metric diagnostics

Author: Marvin Schubert (c) 2025
Version: 1.0.0
License: MIT
"""

from __future__ import annotations
import re
from pathlib import Path
import math
from typing import Dict, List, Tuple, Optional
import argparse

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# --------- Configuration ---------

# Expected metric ordering (kept in German to match upstream export labels)
ORDERED_LABELS = [
    "N",
    "Frametime Ø [ms]",
    "Frametime p95 [ms]",
    "FPS Ø [#]",
    "GPU-Zeit Ø [ms]",
    "GPU-Zeit p95 [ms]",
    "Draw Calls Ø [#]",
    "Primitives Ø [#]",
    "Local VRAM [MB]",
    "Shader Mem [MB]",
]

# Tolerant alias mapping for row labels (robust matching against slight variations)
# key = canonical target label, values = patterns (normalized)
LABEL_ALIASES: Dict[str, List[str]] = {
    "N": ["n"],
    "Frametime Ø [ms]": ["frametimeøms", "frametimeoms", "frametime ms", "frametimeø", "frametimeoems"],
    "Frametime p95 [ms]": ["frametimep95ms", "frametime p95", "p95 ms", "p95frametime"],
    "FPS Ø [#]": ["fpsø#", "fpsoe#", "fps", "fpsø"],
    "GPU-Zeit Ø [ms]": ["gpu-zeitøms", "gpuzeitøms", "gpuzeitoms", "gpuzeit ms", "gpu ø"],
    "GPU-Zeit p95 [ms]": ["gpu-zeitp95ms", "gpuzeitp95ms", "gpu p95"],
    "Draw Calls Ø [#]": ["drawcallsø#", "drawcalls", "draw calls"],
    "Primitives Ø [#]": ["primitivesø#", "primitives", "primitive", "visibleprimitives", "visibleprimitive", "prim"],
    "Local VRAM [MB]": ["localvrammb", "vram", "gpu memory", "gpu mem", "local vram"],
    "Shader Mem [MB]": ["shadermemmb", "shader mem", "shader memory"],
}

# Output formatting precision rules (decimal places)
DECIMALS = {
    "N": 0,
    "FPS Ø [#]": 3,
    "Draw Calls Ø [#]": 0,
    "Primitives Ø [#]": 0,  # integer representation for primitives
    "Local VRAM [MB]": 3,
    "Shader Mem [MB]": 3,
    # default for time metrics (ms): 3
}

# --------- Helpers ---------

def normalize(s: str) -> str:
    """Lowercase and strip non-alphanumerics for robust label matching."""
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def best_label_match(raw_label: str) -> Optional[str]:
    """Map a raw row label to the canonical target label, or None if no match."""
    n = normalize(raw_label)
    for target, patterns in LABEL_ALIASES.items():
        if n == normalize(target):
            return target
        if any(p in n for p in patterns):
            return target
    return None


def parse_number(val) -> Optional[float]:
    """Robust numeric parser handling mixed German/English formatting.

    Logic (adapted from requested update snippet):
      * Accept existing numeric types directly (filter NaN/inf).
      * Treat None / empty / dash as missing.
      * Remove regular and non-breaking spaces.
      * If BOTH '.' and ',' occur: assume German style ('.' thousands, ',' decimal) -> strip '.' then replace ',' with '.'.
      * Else if only ',' occurs: treat it as decimal separator -> replace with '.'.
      * Else if only '.' occurs: treat '.' as decimal separator (leave as is).
      * Else keep digits (e.g. '1000').
      * Final fallback: strip non [0-9 . -] chars and retry.
      * Returns None if still unparsable.
    """
    if val is None:
        return None
    if isinstance(val, (int, float, np.number)):
        try:
            f = float(val)
        except Exception:
            return None
        if math.isnan(f) or math.isinf(f):
            return None
        return f
    s = str(val).strip()
    if s == "" or s == "-":
        return None
    # remove spaces (including NBSP)
    s = s.replace('\xa0', '').replace(' ', '')
    if '.' in s and ',' in s:
        # German style: 1.234,56
        s = s.replace('.', '')
        s = s.replace(',', '.')
    else:
        if ',' in s and '.' not in s:
            # Only comma present -> decimal comma
            s = s.replace(',', '.')
        # else: only '.' or none -> treat '.' as decimal point; if it was a thousands separator we cannot infer safely.
        elif '.' in s and s.count('.') > 1:
            # Multiple dots but no comma -> treat ALL dots as thousands separators
            s = s.replace('.', '')
    try:
        return float(s)
    except ValueError:
        cleaned = re.sub(r'[^0-9.\-]', '', s)
        # If still multiple dots (likely thousands) remove all but last
        if cleaned.count('.') > 1:
            # Keep only the last dot (if interpreting final group as decimals) OR remove all if groups len==0
            parts = cleaned.split('.')
            # Heuristic: if last part length in {1,2,3} and others length 3 -> treat preceding as thousands
            if all(len(p)==3 for p in parts[:-1]) and 1 <= len(parts[-1]) <= 3:
                cleaned = ''.join(parts[:-1]) + '.' + parts[-1]
            else:
                # Fallback: strip all dots -> integer
                cleaned = cleaned.replace('.', '')
        try:
            return float(cleaned)
        except ValueError:
            # Optional warning (could be toggled via env if needed)
            # print(f"WARNING: could not parse number: {s!r}")
            return None

def percent_delta(a: Optional[float], b: Optional[float]) -> float:
    """Compute percentage delta (B vs A) = (B - A)/A * 100.
    Returns NaN if A is None/NaN/0 or B is None/NaN.
    """
    if a is None or b is None:
        return float('nan')
    if isinstance(a, float) and (math.isnan(a) or math.isinf(a)):
        return float('nan')
    if isinstance(b, float) and (math.isnan(b) or math.isinf(b)):
        return float('nan')
    if a == 0:
        return float('nan')
    return (b - a) / a * 100.0

NBSP = "\u00A0"  # non-breaking space for thousands grouping

def _group_int(n: int) -> str:
    s = str(abs(n))
    parts = []
    while s:
        parts.append(s[-3:])
        s = s[:-3]
    grouped = NBSP.join(reversed(parts))
    return f"-{grouped}" if n < 0 else grouped

def fmt_de(label: str, val: Optional[float]) -> str:
    """Format numeric value with decimal comma & NBSP thousands separator.

    Rounding policy:
      - Precision per DECIMALS[label] else 3.
      - Integer metrics (precision 0) displayed without decimal part.
      - Delta column uses label 'Δ' -> 3 decimals.
    """
    if val is None or (isinstance(val, float) and (math.isnan(val) or math.isinf(val))):
        return ""
    # enforce 3 decimals for delta
    if label == "Δ":
        prec = 3
    else:
        prec = DECIMALS.get(label, 3)
    if prec == 0:
        return _group_int(int(round(val)))
    # round and split
    v = round(float(val), prec)
    sign = "-" if v < 0 else ""
    v = abs(v)
    int_part = int(math.floor(v))
    frac_part = v - int_part
    int_str = _group_int(int_part)
    frac_str = f"{frac_part:.{prec}f}".split(".")[1]
    return f"{sign}{int_str},{frac_str}"

def extract_scene_variant(sheet_name: str) -> Tuple[Optional[str], Optional[str]]:
    m = re.search(r"scene\s*(\d+)\s*[_-]\s*([ab])", sheet_name, re.IGNORECASE)
    if m:
        return m.group(1), m.group(2).upper()
    # Fallbacks
    m2 = re.search(r"scene\s*(\d+)", sheet_name, re.IGNORECASE)
    var = "A" if re.search(r"[_-]a\b", sheet_name, re.IGNORECASE) else ("B" if re.search(r"[_-]b\b", sheet_name, re.IGNORECASE) else None)
    return (m2.group(1) if m2 else None), var

# --------- Kernfunktionen ---------

def read_sheet_aggregation(xls_path: Path, sheet: str) -> Tuple[Dict[str, float], Dict[str, List[float]]]:
    """Read one worksheet and compute per-metric mean across run columns.

    Returns:
        agg:  metric label -> mean across runs (NaN if no valid values)
        runs: metric label -> list of individual run values
    """
    df = pd.read_excel(xls_path, sheet_name=sheet, header=None, engine="openpyxl")
    agg: Dict[str, float] = {}
    runs: Dict[str, List[float]] = {}
    for i in range(len(df)):
        raw_label = df.iat[i, 0]
        if not isinstance(raw_label, str):
            continue
        target_label = best_label_match(raw_label)
        if target_label is None:
            continue
        row_vals: List[float] = []
        for j in range(1, df.shape[1]):
            v = parse_number(df.iat[i, j])
            if v is not None:
                row_vals.append(v)
        runs[target_label] = row_vals
        agg[target_label] = float(np.mean(row_vals)) if row_vals else float("nan")
    return agg, runs

def build_comparison_for_scene(scene: str,
                               agg_A: Dict[str, float],
                               agg_B: Dict[str, float]) -> pd.DataFrame:
    """Create comparison dataframe for a scene (Metric | A mean | B mean | Delta [%])."""
    rows = []
    for label in ORDERED_LABELS:
        a = agg_A.get(label, float("nan"))
        b = agg_B.get(label, float("nan"))
        delta = percent_delta(a, b)
        rows.append({
            "Kennzahl": label,
            "A (Ø)": a,
            "B (Ø)": b,
            "Δ B vs A [%]": delta
        })
    return pd.DataFrame(rows, columns=["Kennzahl", "A (Ø)", "B (Ø)", "Δ B vs A [%]"])

def write_scene_sheets(wb: Workbook,
                       scene: str,
                       agg_A: Dict[str, float],
                       agg_B: Dict[str, float],
                       cmp_df: pd.DataFrame) -> None:
    """Write three sheets per scene: aggregated A, aggregated B, and comparison."""
    def _write_agg(sheet_title: str, agg: Dict[str, float], variant: str):
        ws = wb.create_sheet(title=sheet_title)
        ws["A1"] = f"Aggregated metrics – Variant {variant}"
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="left")

        ws["A2"] = "Metric"
        ws["B2"] = "Mean over runs"
        ws["A2"].font = ws["B2"].font = Font(bold=True)
        ws["A2"].alignment = ws["B2"].alignment = Alignment(horizontal="center")

        r = 3
        for label in ORDERED_LABELS:
            ws.cell(row=r, column=1, value=label)
            val = agg.get(label, float("nan"))
            ws.cell(row=r, column=2, value=fmt_de(label, val))
            r += 1

        # Breiten
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 16

    # Agg A
    _write_agg(f"Scene{scene}_Agg_A", agg_A, "A")
    # Agg B
    _write_agg(f"Scene{scene}_Agg_B", agg_B, "B")

    # Comparison sheet with methodology note
    ws = wb.create_sheet(title=f"Scene{scene}_Vergleich")
    note_text = ("Δ = (B − A) / A · 100. Rounding: time & memory metrics 3 decimals; FPS 3 decimals; "
                 "integer counters (N, Draw Calls, Primitives) 0 decimals; Δ 3 decimals. Decimal comma, NBSP thousands.")
    ws.cell(row=1, column=1, value=note_text)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
    headers = ["Metric", "A (Ø)", "B (Ø)", "Δ B vs A [%]"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    # Write data rows starting at row 3
    for excel_row, (_, row) in enumerate(cmp_df.iterrows(), start=3):
        lab = row["Kennzahl"]
        ws.cell(row=excel_row, column=1, value=lab)
        ws.cell(row=excel_row, column=2, value=fmt_de(lab, row["A (Ø)"]))
        ws.cell(row=excel_row, column=3, value=fmt_de(lab, row["B (Ø)"]))
        ws.cell(row=excel_row, column=4, value=fmt_de("Δ", row["Δ B vs A [%]"]))

    # Spaltenbreiten:
    widths = [24, 16, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _collect_from_single_file(path: Path) -> List[Dict[str, object]]:
    """Collect A/B variant sheets from a single workbook (legacy mode)."""
    xls = pd.ExcelFile(path, engine="openpyxl")
    recs: List[Dict[str, object]] = []
    for sheet in xls.sheet_names:
        scene, var = extract_scene_variant(sheet)
        if scene is None or var not in ("A", "B"):
            continue
        agg, runs = read_sheet_aggregation(path, sheet)
        recs.append({
            "scene": str(scene),
            "variant": var,
            "agg": agg,
            "runs": runs,
            "sheet": sheet,
            "source": path.name,
        })
    return recs

def _collect_from_two_files(path_a: Path, path_b: Path) -> List[Dict[str, object]]:
    """Collect scene data from two workbooks (one per variant)."""
    def collect_one(p: Path, variant: str) -> List[Dict[str, object]]:
        if not p.exists():
            raise FileNotFoundError(f"Datei nicht gefunden: {p}")
        xls = pd.ExcelFile(p, engine="openpyxl")
        out: List[Dict[str, object]] = []
        for sheet in xls.sheet_names:
            scene, var_in_sheet = extract_scene_variant(sheet)
            # If sheet lacks variant suffix we assign the variant from the file context
            if scene is None:
                # Versuche einfache Szeneextraktion ohne Variantensuffix
                m = re.search(r"scene\s*(\d+)", sheet, re.IGNORECASE)
                if m:
                    scene = m.group(1)
            if scene is None:
                continue
            effective_variant = variant if var_in_sheet is None else var_in_sheet
            # Enforce passed variant if embedded suffix disagrees
            if effective_variant != variant:
                effective_variant = variant
            agg, runs = read_sheet_aggregation(p, sheet)
            out.append({
                "scene": str(scene),
                "variant": effective_variant,
                "agg": agg,
                "runs": runs,
                "sheet": sheet,
                "source": p.name,
            })
        return out
    recs_a = collect_one(path_a, "A")
    recs_b = collect_one(path_b, "B")
    return recs_a + recs_b

def main():
    parser = argparse.ArgumentParser(description="A/B comparison for Unreal Engine CSV metric aggregations")
    parser.add_argument("--a", dest="file_a", type=Path, help="Workbook for variant A", required=False)
    parser.add_argument("--b", dest="file_b", type=Path, help="Workbook for variant B", required=False)
    parser.add_argument("--auto", action="store_true", help="Auto-detect 'messungen_auswertung_a.xlsx' and '_b.xlsx' if present")
    parser.add_argument("--out", dest="out", type=Path, default=Path("messungen_ab_vergleich.xlsx"), help="Output workbook path")
    parser.add_argument("--no-recompute-fps", dest="recompute_fps", action="store_false", help="Do NOT recompute FPS from frametime means")
    parser.add_argument("--debug", action="store_true", help="Print debug info about parsed sheets and metrics")
    parser.set_defaults(recompute_fps=True)
    args = parser.parse_args()

    records: List[Dict[str, object]] = []

    single_default = Path("messungen_auswertung.xlsx")
    default_a = Path("messungen_auswertung_a.xlsx")
    default_b = Path("messungen_auswertung_b.xlsx")

    mode = "single"

    if args.file_a or args.file_b:
        if not (args.file_a and args.file_b):
            raise SystemExit("--a and --b must be provided together.")
        records = _collect_from_two_files(args.file_a, args.file_b)
        mode = "two-files-cli"
    elif args.auto and default_a.exists() and default_b.exists():
        records = _collect_from_two_files(default_a, default_b)
        mode = "two-files-auto"
    elif default_a.exists() and default_b.exists():
        # Auto-Erkennung ohne --auto, falls beide exakt so heißen
        records = _collect_from_two_files(default_a, default_b)
        mode = "two-files-autodetect"
    else:
        if not single_default.exists():
            raise FileNotFoundError("No input workbooks found: expected either 'messungen_auswertung.xlsx' OR both 'messungen_auswertung_a.xlsx' & 'messungen_auswertung_b.xlsx'.")
        records = _collect_from_single_file(single_default)
        mode = "single"

    if not records:
        raise RuntimeError("No matching sheets found (Scene{n}_A / Scene{n}_B or Scene{n} in A/B files).")

    # Szenen-Liste
    scenes = sorted(sorted({r["scene"] for r in records}), key=lambda s: int(re.findall(r"\d+", s)[0]))
    by_scene_var: Dict[Tuple[str, str], Dict[str, float]] = {}
    by_scene_var_runs: Dict[Tuple[str, str], Dict[str, List[float]]] = {}
    # Optional FPS-Recompute (aus Frametime je Lauf)
    if args.recompute_fps:
        for r in records:
            runs = r.get("runs", {})  # type: ignore
            ft_runs = runs.get("Frametime Ø [ms]", [])  # type: ignore
            if ft_runs:
                fps_runs = [1000.0 / v for v in ft_runs if isinstance(v, (int, float)) and v > 0]
                if fps_runs:
                    new_fps = float(np.mean(fps_runs))
                    orig_fps = r["agg"].get("FPS Ø [#]")  # type: ignore
                    r["agg"]["FPS Ø [#]"] = new_fps  # type: ignore
                    r["_fps_diff"] = (orig_fps, new_fps)  # type: ignore
    for r in records:
        by_scene_var[(r["scene"], r["variant"])] = r["agg"]  # type: ignore
        by_scene_var_runs[(r["scene"], r["variant"])] = r.get("runs", {})  # type: ignore

    if args.debug:
        print("--- DEBUG: Parsed Sheets Summary ---")
        for r in records:
            scene = r["scene"]; var = r["variant"]
            runs = r.get("runs", {})  # type: ignore
            print(f"Scene {scene} Variant {var} (sheet={r['sheet']}, source={r['source']}):")
            for lab, vals in runs.items():
                print(f"  - {lab:22s} runs={len(vals)} sample={vals[:3]}")
        print("--- END DEBUG ---")

    wb = Workbook()
    if wb.active and wb.active.title == "Sheet":
        wb.remove(wb.active)

    ws_sum = wb.create_sheet(title="Gesamtübersicht")  # Keeping German sheet name for continuity
    ws_sum.append(["Scene", "Metric", "A (Ø)", "B (Ø)", "Δ B vs A [%]"])
    for c in range(1, 6):
        ws_sum.cell(row=1, column=c).font = Font(bold=True)
        ws_sum.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    sum_row = 2
    warn_incomplete: List[str] = []
    for scene in scenes:
        agg_A = by_scene_var.get((scene, "A"), {})
        agg_B = by_scene_var.get((scene, "B"), {})
        if not agg_A or not agg_B:
            warn_incomplete.append(scene)
        cmp_df = build_comparison_for_scene(scene, agg_A, agg_B)
        # Warn if metric missing for one variant but present for the other
        if args.debug:
            for lab in ORDERED_LABELS:
                a_vals = by_scene_var_runs.get((scene, "A"), {}).get(lab, [])
                b_vals = by_scene_var_runs.get((scene, "B"), {}).get(lab, [])
                if (not b_vals and a_vals) or (not a_vals and b_vals):
                    print(f"DEBUG WARN: Scene{scene} metric '{lab}' missing values for variant {'B' if (not b_vals and a_vals) else 'A'}")
        write_scene_sheets(wb, scene, agg_A, agg_B, cmp_df)
        for _, row in cmp_df.iterrows():
            ws_sum.cell(row=sum_row, column=1, value=f"Scene{scene}")
            lab = row["Kennzahl"]  # original German label retained
            ws_sum.cell(row=sum_row, column=2, value=lab)
            ws_sum.cell(row=sum_row, column=3, value=fmt_de(lab, row["A (Ø)"]))
            ws_sum.cell(row=sum_row, column=4, value=fmt_de(lab, row["B (Ø)"]))
            ws_sum.cell(row=sum_row, column=5, value=fmt_de("Δ", row["Δ B vs A [%]"]))
            sum_row += 1
        sum_row += 1

    for col, w in zip(range(1, 6), [14, 24, 16, 16, 16]):
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    out_path = args.out
    wb.save(out_path)
    msg = f"✓ A/B comparison exported ({mode}): {out_path.resolve()}"
    if args.recompute_fps:
        diffs = [r for r in records if r.get("_fps_diff") and all(isinstance(x, (int,float)) for x in r.get("_fps_diff"))]
        if diffs:
            # Prüfe, ob Differenz signifikant (>0.1 fps)
            changed = []
            for r in diffs:
                orig, new = r["_fps_diff"]  # type: ignore
                if orig is not None and abs(orig - new) > 0.1:
                    changed.append(r["scene"]+r["variant"])  # type: ignore
            if changed:
                msg += f" | FPS recomputed from frametime: deviation >0.1 at {', '.join(changed)}"
    if warn_incomplete:
        msg += f" | Incomplete scenes (only one variant): {', '.join('Scene'+s for s in warn_incomplete)}"
    print(msg)

if __name__ == "__main__":
    main()

